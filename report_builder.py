import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C_DARK="1F3864"; C_MID="2F5496"; C_ACCENT="4472C4"
C_GREEN="E2EFDA"; C_RED="FCE4D6"; C_STRIPE="EBF3FF"
C_WHITE="FFFFFF"; C_YELLOW="FFF2CC"
CITY_COLORS={"ORD":"7030A0","SDF":"0070C0","MEM":"FF9900","BTR":"C00000","SHV":"00B050"}

def bdr():
    s=Side(style='thin',color="CCCCCC")
    return Border(left=s,right=s,top=s,bottom=s)

def hdr(cell,bg=C_DARK):
    cell.font=Font(name='Arial',bold=True,color="FFFFFF",size=11)
    cell.fill=PatternFill("solid",fgColor=bg)
    cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    cell.border=bdr()

def dat(cell,bg=C_WHITE,bold=False,align='right',fmt=None):
    cell.font=Font(name='Arial',bold=bold,size=10)
    cell.fill=PatternFill("solid",fgColor=bg)
    cell.alignment=Alignment(horizontal=align,vertical='center')
    cell.border=bdr()
    if fmt: cell.number_format=fmt

def title_row(ws, text, cols, bg, size=13):
    ws.merge_cells(f'A1:{get_column_letter(cols)}1')
    ws['A1']=text
    ws['A1'].font=Font(name='Arial',bold=True,size=size,color="FFFFFF")
    ws['A1'].fill=PatternFill("solid",fgColor=bg)
    ws['A1'].alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height=32

def build_report(city_results, out_path):
    wb=openpyxl.Workbook()
    wb.remove(wb.active)

    # ══ RESUMEN GENERAL ═══════════════════════════════════════════════════════
    ws=wb.create_sheet("📊 Resumen General")
    ws.sheet_view.showGridLines=False
    for col,w in zip('ABCDEFG',[24,16,14,14,14,10,10]):
        ws.column_dimensions[col].width=w

    period=list(city_results.values())[0]['period']
    title_row(ws,f"REPORTE SEMANAL — TODAS LAS CIUDADES — {period}",7,C_DARK,14)

    ws.row_dimensions[3].height=26
    for ci,h in enumerate(['Ciudad','Revenue GOFO','Payroll','Profit','Margen','Drivers'],1):
        hdr(ws.cell(row=3,column=ci,value=h),bg=C_ACCENT)

    row=4; totals={'rev':0,'pay':0,'prof':0,'drv':0}
    for code,res in city_results.items():
        ws.row_dimensions[row].height=22
        cc=CITY_COLORS.get(code,C_ACCENT)
        bg=C_STRIPE if row%2==0 else C_WHITE
        c=ws.cell(row=row,column=1,value=f"{code} – {res['city_name']}")
        c.font=Font(name='Arial',bold=True,size=10,color="FFFFFF")
        c.fill=PatternFill("solid",fgColor=cc)
        c.alignment=Alignment(horizontal='left',vertical='center')
        c.border=bdr()
        profit_bg=C_GREEN if res['profit']>=0 else C_RED
        for ci,(val,fmt,rbg) in enumerate([
            (res['revenue'],'"$"#,##0.00',bg),
            (res['total_payroll'],'"$"#,##0.00',bg),
            (res['profit'],'"$"#,##0.00',profit_bg),
            (res['margin']/100,'0.0%',profit_bg),
            (len(res['drivers_list']),'#,##0',bg),
        ],2):
            c2=ws.cell(row=row,column=ci,value=val)
            dat(c2,bg=rbg,align='right' if ci<6 else 'center',fmt=fmt)
        totals['rev']+=res['revenue']; totals['pay']+=res['total_payroll']
        totals['prof']+=res['profit']; totals['drv']+=len(res['drivers_list'])
        row+=1

    ws.row_dimensions[row+1].height=26
    tm=(totals['prof']/totals['rev']*100) if totals['rev']>0 else 0
    for ci,(val,fmt) in enumerate([
        ("TOTAL SEMANA",None),(totals['rev'],'"$"#,##0.00'),
        (totals['pay'],'"$"#,##0.00'),(totals['prof'],'"$"#,##0.00'),
        (tm/100,'0.0%'),(totals['drv'],'#,##0')],1):
        c=ws.cell(row=row+1,column=ci,value=val)
        c.font=Font(name='Arial',bold=True,color="FFFFFF",size=11)
        c.fill=PatternFill("solid",fgColor=C_DARK)
        c.alignment=Alignment(horizontal='center' if ci==1 else 'right',vertical='center')
        c.border=bdr()
        if fmt: c.number_format=fmt

    # ══ HOJAS POR CIUDAD ══════════════════════════════════════════════════════
    for code,res in city_results.items():
        cc=CITY_COLORS.get(code,C_ACCENT)
        drivers=res['drivers_list']

        # ── PAYROLL ──────────────────────────────────────────────────────────
        ws_p=wb.create_sheet(f"💵 {code} Payroll")
        ws_p.sheet_view.showGridLines=False
        cols_p=[('Driver',24),('Rutas',22),('Stops',8),('Pago Stops',13),
                ('Extras',8),('Pago Extra',12),('Bono Fijo',11),
                ('Bono Extra',11),('Cobro Extra',11),('Penaliz.',11),('TOTAL',13)]
        for ci,(_, w) in enumerate(cols_p,1):
            ws_p.column_dimensions[get_column_letter(ci)].width=w
        title_row(ws_p,f"PAYROLL — {code} {res['city_name']} — {res['period']}",len(cols_p),cc)
        ws_p.row_dimensions[2].height=26
        for ci,(h,_) in enumerate(cols_p,1):
            hdr(ws_p.cell(row=2,column=ci,value=h),bg=cc)

        for ri,drv in enumerate(drivers,3):
            bg=C_STRIPE if ri%2==0 else C_WHITE
            ws_p.row_dimensions[ri].height=20
            stops_t=sum(r['stops']  for r in drv['rutas_data'])
            ext_t  =sum(r['extras'] for r in drv['rutas_data'])
            nota=f" [{drv['ajuste_nota']}]" if drv.get('ajuste_nota') else ""
            row_vals=[
                (drv['driver'],          'left',  None,           bg,      False),
                (drv['rutas']+nota,      'left',  None,           bg,      False),
                (stops_t,                'center','#,##0',         bg,      False),
                (drv['pago_base'] - sum(r['pago_extra'] for r in drv['rutas_data']),
                                         'right', '"$"#,##0.00',  bg,      False),
                (ext_t,                  'center','#,##0',         bg,      False),
                (sum(r['pago_extra'] for r in drv['rutas_data']),
                                         'right', '"$"#,##0.00',  bg,      False),
                (drv['bono_fijo'],        'right', '"$"#,##0.00',  C_YELLOW if drv['bono_fijo']>0 else bg, False),
                (drv.get('ajuste_bono',0),'right','"$"#,##0.00',  C_YELLOW if drv.get('ajuste_bono',0)>0 else bg, False),
                (drv.get('ajuste_cobro',0),'right','"$"#,##0.00', C_RED    if drv.get('ajuste_cobro',0)>0 else bg, False),
                (drv['penalizaciones'],   'right', '"$"#,##0.00',  C_RED    if drv['penalizaciones']<0 else bg, False),
                (drv['pago_total'],       'right', '"$"#,##0.00',  C_GREEN, True),
            ]
            for ci,(val,align,fmt,rbg,bold) in enumerate(row_vals,1):
                c=ws_p.cell(row=ri,column=ci,value=val)
                dat(c,bg=rbg,bold=bold,align=align,fmt=fmt)

        tr=len(drivers)+3
        ws_p.row_dimensions[tr].height=24
        ws_p.merge_cells(f'A{tr}:I{tr}')
        tc=ws_p[f'A{tr}']
        tc.value="TOTALES"
        tc.font=Font(name='Arial',bold=True,color="FFFFFF",size=11)
        tc.fill=PatternFill("solid",fgColor=cc)
        tc.alignment=Alignment(horizontal='center',vertical='center')
        tc.border=bdr()
        for ci,val in [(10,sum(d['penalizaciones'] for d in drivers)),(11,sum(d['pago_total'] for d in drivers))]:
            c=ws_p.cell(row=tr,column=ci,value=val)
            c.font=Font(name='Arial',bold=True,color="FFFFFF",size=11)
            c.fill=PatternFill("solid",fgColor=cc)
            c.alignment=Alignment(horizontal='right',vertical='center')
            c.border=bdr()
            c.number_format='"$"#,##0.00'

        # ── PROFIT ────────────────────────────────────────────────────────────
        ws_r=wb.create_sheet(f"📈 {code} Profit")
        ws_r.sheet_view.showGridLines=False
        cols_r=[('Driver',24),('Rutas',22),('Revenue GOFO',15),('Pago Driver',14),('Profit',13),('Margen %',11)]
        for ci,(_,w) in enumerate(cols_r,1):
            ws_r.column_dimensions[get_column_letter(ci)].width=w
        title_row(ws_r,f"REVENUE Y PROFIT — {code} {res['city_name']}",len(cols_r),cc)
        ws_r.row_dimensions[2].height=26
        for ci,(h,_) in enumerate(cols_r,1):
            hdr(ws_r.cell(row=2,column=ci,value=h),bg=cc)

        for ri,drv in enumerate(drivers,3):
            bg=C_STRIPE if ri%2==0 else C_WHITE
            ws_r.row_dimensions[ri].height=20
            pbg=C_GREEN if drv['profit']>=0 else C_RED
            for ci,(val,align,fmt,rbg,bold) in enumerate([
                (drv['driver'],    'left', None,           bg,   False),
                (drv['rutas'],     'left', None,           bg,   False),
                (drv['revenue'],   'right','"$"#,##0.00',  bg,   False),
                (drv['pago_total'],'right','"$"#,##0.00',  bg,   False),
                (drv['profit'],    'right','"$"#,##0.00',  pbg,  True),
                (drv['margin']/100,'center','0.0%',         pbg,  False),
            ],1):
                c=ws_r.cell(row=ri,column=ci,value=val)
                dat(c,bg=rbg,bold=bold,align=align,fmt=fmt)

        tr2=len(drivers)+3
        ws_r.row_dimensions[tr2].height=24
        tv=sum(d['revenue'] for d in drivers)
        tp=sum(d['pago_total'] for d in drivers)
        tpf=sum(d['profit'] for d in drivers)
        tm2=(tpf/tv*100) if tv>0 else 0
        ws_r.merge_cells(f'A{tr2}:B{tr2}')
        tc2=ws_r[f'A{tr2}']
        tc2.value="TOTALES"; tc2.font=Font(name='Arial',bold=True,color="FFFFFF",size=11)
        tc2.fill=PatternFill("solid",fgColor=cc)
        tc2.alignment=Alignment(horizontal='center',vertical='center'); tc2.border=bdr()
        for ci,val,fmt in [(3,tv,'"$"#,##0.00'),(4,tp,'"$"#,##0.00'),(5,tpf,'"$"#,##0.00'),(6,tm2/100,'0.0%')]:
            c=ws_r.cell(row=tr2,column=ci,value=val)
            c.font=Font(name='Arial',bold=True,color="FFFFFF",size=11)
            c.fill=PatternFill("solid",fgColor=cc)
            c.alignment=Alignment(horizontal='right' if ci<6 else 'center',vertical='center')
            c.border=bdr(); c.number_format=fmt

    wb.save(out_path)
    return out_path
