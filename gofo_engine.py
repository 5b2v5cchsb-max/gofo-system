import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os, sys, glob

# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIGURACIÓN DE TARIFAS
# ═══════════════════════════════════════════════════════════════════════════════
CITY_CONFIG = {
    "ORD": {
        "name": "Chicago",
        "routes": {
            "ORD01-032": {"stop": 1.60, "extra": 0.25},
            "ORD01-007": {"stop": 1.70, "extra": 0.40},
        },
        "default_route": "ORD01-032",
        "bonos_fijos": {"susy": 800.0},
    },
    "SDF": {
        "name": "Louisville",
        "routes": {
            "SDF01-014": {"stop": 1.60, "extra": 0.50},
            "SDF01-027": {"stop": 1.70, "extra": 0.50},
        },
        "default_route": "SDF01-014",
        "bonos_fijos": {"yadrian": 800.0},
    },
    "MEM": {
        "name": "Memphis",
        "routes": {
            "MEM01-014": {"stop": 1.60, "extra": 0.40},
            "MEM01-020": {"stop": 1.79, "extra": 0.40},
            "MEM01-024": {"stop": 1.89, "extra": 0.40},
        },
        "default_route": "MEM01-014",
        "bonos_fijos": {"areli": 500.0},
    },
    "BTR": {
        "name": "Baton Rouge",
        "routes": {
            "BTR01-007": {"stop": 1.50, "extra": 0.25},
        },
        "default_route": "BTR01-007",
        "bonos_fijos": {"heather": 500.0},
    },
    "SHV": {
        "name": "Shreveport",
        "routes": {
            "SHV01-002": {"stop": 1.50, "extra": 0.25},
        },
        "default_route": "SHV01-002",
        "bonos_fijos": {"germai": 500.0},
    },
}

C_DARK="1F3864"; C_MID="2F5496"; C_ACCENT="4472C4"
C_GREEN="E2EFDA"; C_RED="FCE4D6"; C_STRIPE="EBF3FF"
C_WHITE="FFFFFF"; C_YELLOW="FFF2CC"
CITY_COLORS={"ORD":"7030A0","SDF":"0070C0","MEM":"FF9900","BTR":"C00000","SHV":"00B050"}

def bdr():
    s=Side(style='thin',color="CCCCCC")
    return Border(left=s,right=s,top=s,bottom=s)

def hdr(cell,bg=C_DARK):
    cell.font=Font(name='Arial',bold=True,color="FFFFFF",size=11)
    cell.fill=PatternFill("solid",fgColor=bg)
    cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    cell.border=bdr()

def dat(cell,bg=C_WHITE,bold=False,align='right',fmt=None):
    cell.font=Font(name='Arial',bold=bold,size=10)
    cell.fill=PatternFill("solid",fgColor=bg)
    cell.alignment=Alignment(horizontal=align,vertical='center')
    cell.border=bdr()
    if fmt: cell.number_format=fmt

def detect_city(filepath):
    fname = os.path.basename(filepath).upper()
    for code in CITY_CONFIG:
        if fname.startswith(code + "-"):
            return code
    return None

# ─── PROCESAR UN ARCHIVO ──────────────────────────────────────────────────────
def process_file(filepath):
    city_code = detect_city(filepath)
    cfg = CITY_CONFIG[city_code]
    xl  = pd.ExcelFile(filepath)

    # ── DSP Summary ──────────────────────────────────────────────────────────
    dsp = pd.read_excel(xl, sheet_name='DSP Summary', header=None).iloc[2]
    # col6 = Total Billing Amount = what GOFO deposits (already minus claims)
    # col7 = Claim Amount (negative) = already deducted from col6
    # col10 = Delivery fee gross (col6 + abs(col7))
    revenue_total = float(dsp[6])   # ← what GOFO actually pays you
    claims_total  = float(dsp[7])   # ← for reference only
    period        = str(dsp[4])

    # ── Driver Summary — gross fee per driver ────────────────────────────────
    ds_raw = pd.read_excel(xl, sheet_name='Driver Summary', header=None)
    driver_fee = {}   # col6 per driver = gross delivery fee
    for _, r in ds_raw.iloc[2:].iterrows():
        name = str(r[1]).strip()
        if not name or name == 'nan': continue
        if any(x in name.lower() for x in ['payroll','gofo','profit','total']): break
        driver_fee[name] = float(r[6]) if pd.notna(r[6]) else 0.0

    # ── Details — paquete a paquete ──────────────────────────────────────────
    det_raw = pd.read_excel(xl, sheet_name='Details of Delivery Fees', header=None)
    det = det_raw.iloc[1:].copy()
    det.columns = det_raw.iloc[0].values
    det = det[det["Courier"].notna()].copy()
    det['Courier'] = det['Courier'].astype(str).str.strip()  # fix trailing spaces
    det['STOP Point Details'] = pd.to_numeric(det['STOP Point Details'], errors='coerce').fillna(1)
    det['Region/route'] = det['Region/route'].astype(str).str.strip()

    # ── Claims por driver (para descontar del payroll) ───────────────────────
    try:
        cl = pd.read_excel(xl, sheet_name='Claims Detail')
        cl['amt'] = pd.to_numeric(cl['The total expenses-exclusive of taxes'], errors='coerce').fillna(0)
        penalties = cl[cl['amt'] < 0].copy()
    except:
        penalties = pd.DataFrame(columns=['Courier','amt','Claim Type','Waybill No.','Remarks'])

    try:
        off_raw = pd.read_excel(xl, sheet_name='Offset Details', header=None)
        off = off_raw.iloc[1:].copy()
        off.columns = off_raw.iloc[0].values
        off = off[off['Courier'].notna()].copy()
        off['amt'] = pd.to_numeric(off['The total expenses-exclusive of taxes'], errors='coerce').fillna(0)
        off_neg = off[off['amt'] < 0].copy()
    except:
        off_neg = pd.DataFrame(columns=['Courier','amt'])

    # ── CALCULAR PAYROLL POR DRIVER ──────────────────────────────────────────
    driver_rows = []
    detail_rows = []

    for driver, gross_fee in driver_fee.items():
        d_det = det[det['Courier'] == driver]
        if len(d_det) == 0: continue

        # Bono fijo
        bono = 0.0
        for key, val in cfg['bonos_fijos'].items():
            if key.lower() in driver.lower():
                bono = val

        # Penalizaciones del driver (se descuentan del payroll)
        pen     = float(penalties[penalties['Courier'] == driver]['amt'].sum()) if len(penalties) > 0 else 0.0
        off_pen = 0.0  # offsets not charged to driver (small retroactive adjustments)
        total_pen = pen  # only claims, not offsets   # always negative or zero
        n_pen   = int(len(penalties[penalties['Courier'] == driver])) if len(penalties) > 0 else 0

        # Stops y extras por ruta
        pago_stops_total = 0.0
        pago_extra_total = 0.0
        rutas_trabajadas = []

        for ruta, grp in d_det.groupby('Region/route'):
            route_cfg    = cfg['routes'].get(ruta, cfg['routes'][cfg['default_route']])
            tarifa_stop  = route_cfg['stop']
            tarifa_extra = route_cfg['extra']
            stops  = int((grp['STOP Point Details'] == 1).sum())
            extras = int((grp['STOP Point Details'] > 1).sum())
            ps = round(stops  * tarifa_stop,  2)
            pe = round(extras * tarifa_extra, 2)
            pago_stops_total += ps
            pago_extra_total += pe
            rutas_trabajadas.append(ruta)
            detail_rows.append({
                'city': city_code, 'driver': driver,
                'ruta': ruta, 'tarifa_stop': tarifa_stop, 'tarifa_extra': tarifa_extra,
                'stops': stops, 'extras': extras,
                'pago_stops': ps, 'pago_extra': pe,
            })

        # Pago total al driver
        # = pago por stops + pago por extras + bono fijo + penalizaciones
        # penalizaciones son negativas → reducen el pago
        pago_total = round(pago_stops_total + pago_extra_total + bono + total_pen, 2)

        # Revenue de este driver = su gross fee (col6)
        # Profit por driver = gross_fee - pago_total
        # Nota: GOFO descuenta claims del total, pero tú los recuperas del driver
        # entonces el profit neto es el mismo que si usaras gross
        profit = round(gross_fee - pago_total, 2)
        margin = round((profit / gross_fee * 100) if gross_fee > 0 else 0, 1)

        driver_rows.append({
            'city': city_code, 'city_name': cfg['name'],
            'driver': driver,
            'rutas': " + ".join(sorted(set(rutas_trabajadas))),
            'rutas_data': [r for r in detail_rows if r['driver']==driver and r['city']==city_code],
            'gross_fee': gross_fee,
            'pago_stops': round(pago_stops_total, 2),
            'pago_extra': round(pago_extra_total, 2),
            'pago_base':  round(pago_stops_total + pago_extra_total, 2),
            'bono': bono,
            'penalizaciones': round(total_pen, 2),
            'pago_total': pago_total,
            'profit': profit,
            'margin': margin,
            'n_penalizaciones': n_pen,
            'ajuste_bono': 0.0,
            'ajuste_cobro': 0.0,
            'ajuste_nota': '',
        })

    df_drivers = pd.DataFrame(driver_rows)
    total_payroll = round(df_drivers['pago_total'].sum(), 2)

    # ── PROFIT TOTAL ─────────────────────────────────────────────────────────
    # Revenue = what GOFO deposits (col6 DSP = already minus claims)
    # Payroll = what you pay drivers (already recovered claims from them)
    # Profit  = Revenue - Payroll
    # The claims are a wash: GOFO took them from you, you took them from driver
    profit_total = round(revenue_total - total_payroll, 2)
    margin_total = round((profit_total / revenue_total * 100) if revenue_total > 0 else 0, 1)

    return {
        'city': city_code, 'city_name': cfg['name'], 'period': period,
        'revenue': revenue_total,        # DSP col6 = net GOFO deposits
        'claims':  claims_total,         # for info only
        'gross_before_claims': round(revenue_total - claims_total, 2),  # col10
        'total_payroll': total_payroll,
        'profit': profit_total,
        'margin': margin_total,
        'drivers': df_drivers,
        'detail': pd.DataFrame(detail_rows),
        'claims_df': penalties,
    }

# ─── CONSTRUIR EXCEL ──────────────────────────────────────────────────────────
def build_report(city_results, out_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ══ RESUMEN GENERAL ═══════════════════════════════════════════════════════
    ws = wb.create_sheet("📊 Resumen General")
    ws.sheet_view.showGridLines = False
    for col, w in zip('ABCDEFGH', [24,16,14,14,14,10,10,2]):
        ws.column_dimensions[col].width = w

    period = list(city_results.values())[0]['period']
    ws.merge_cells('A1:G1')
    ws['A1'] = f"REPORTE SEMANAL — TODAS LAS CIUDADES — {period}"
    ws['A1'].font = Font(name='Arial', bold=True, size=15, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor=C_DARK)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:G2')
    ws['A2'] = "Revenue = lo que GOFO deposita (neto)  |  Profit = Revenue − Payroll"
    ws['A2'].font = Font(name='Arial', size=10, color="FFFFFF")
    ws['A2'].fill = PatternFill("solid", fgColor=C_MID)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[4].height = 28
    for ci, h in enumerate(['Ciudad','Revenue (neto GOFO)','Claims (info)','Payroll','Profit','Margen','Drivers'], 1):
        hdr(ws.cell(row=4, column=ci, value=h), bg=C_ACCENT)

    row = 5
    totals = {'rev':0,'cl':0,'pay':0,'prof':0,'drv':0}
    for code, res in city_results.items():
        ws.row_dimensions[row].height = 22
        cc = CITY_COLORS.get(code, C_ACCENT)
        bg = C_STRIPE if row % 2 == 0 else C_WHITE
        c = ws.cell(row=row, column=1, value=f"{code} – {res['city_name']}")
        c.font = Font(name='Arial', bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=cc)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = bdr()
        profit_bg = C_GREEN if res['profit'] >= 0 else C_RED
        for ci, (val, fmt, rbg) in enumerate([
            (res['revenue'],       '"$"#,##0.00', bg),
            (res['claims'],        '"$"#,##0.00', C_RED if res['claims'] < 0 else bg),
            (res['total_payroll'], '"$"#,##0.00', bg),
            (res['profit'],        '"$"#,##0.00', profit_bg),
            (res['margin']/100,    '0.0%',         profit_bg),
            (len(res['drivers']),  '#,##0',        bg),
        ], 2):
            c2 = ws.cell(row=row, column=ci, value=val)
            dat(c2, bg=rbg, align='right' if ci < 7 else 'center', fmt=fmt)
        totals['rev']  += res['revenue'];  totals['cl']  += res['claims']
        totals['pay']  += res['total_payroll']
        totals['prof'] += res['profit'];   totals['drv'] += len(res['drivers'])
        row += 1

    ws.row_dimensions[row+1].height = 26
    tm = (totals['prof']/totals['rev']*100) if totals['rev'] > 0 else 0
    for ci, (val, fmt) in enumerate([
        ("TOTAL SEMANA", None),
        (totals['rev'],  '"$"#,##0.00'), (totals['cl'],  '"$"#,##0.00'),
        (totals['pay'],  '"$"#,##0.00'), (totals['prof'], '"$"#,##0.00'),
        (tm/100,         '0.0%'),        (totals['drv'],  '#,##0'),
    ], 1):
        c = ws.cell(row=row+1, column=ci, value=val)
        c.font = Font(name='Arial', bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=C_DARK)
        c.alignment = Alignment(horizontal='center' if ci==1 else 'right', vertical='center')
        c.border = bdr()
        if fmt: c.number_format = fmt

    # ══ HOJAS POR CIUDAD ══════════════════════════════════════════════════════
    for code, res in city_results.items():
        df  = res['drivers']
        cc  = CITY_COLORS.get(code, C_ACCENT)

        # ── PAYROLL ──────────────────────────────────────────────────────────
        ws_p = wb.create_sheet(f"💵 {code} Payroll")
        ws_p.sheet_view.showGridLines = False
        cols_p = [('Driver',24),('Rutas',20),('Stops',8),('Pago Stops',13),
                  ('Extras',8),('Pago Extra',12),('Bono',10),
                  ('Penaliz.',11),('TOTAL PAGO',13)]
        for ci, (_, w) in enumerate(cols_p, 1):
            ws_p.column_dimensions[get_column_letter(ci)].width = w

        ws_p.merge_cells(f'A1:{get_column_letter(len(cols_p))}1')
        ws_p['A1'] = f"PAYROLL — {code} {res['city_name']} — {res['period']}"
        ws_p['A1'].font = Font(name='Arial', bold=True, size=13, color="FFFFFF")
        ws_p['A1'].fill = PatternFill("solid", fgColor=cc)
        ws_p['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_p.row_dimensions[1].height = 30

        ws_p.row_dimensions[2].height = 26
        for ci, (h, _) in enumerate(cols_p, 1):
            hdr(ws_p.cell(row=2, column=ci, value=h), bg=cc)

        for ri, drv in enumerate(df.itertuples(), 3):
            bg = C_STRIPE if ri % 2 == 0 else C_WHITE
            ws_p.row_dimensions[ri].height = 20
            stops_t  = sum(r['stops']  for r in drv.rutas_data) if isinstance(drv.rutas_data, list) else 0
            extras_t = sum(r['extras'] for r in drv.rutas_data) if isinstance(drv.rutas_data, list) else 0
            row_vals = [
                (drv.driver,         'left',  None,           bg,      False),
                (drv.rutas,          'left',  None,           bg,      False),
                (stops_t,            'center','#,##0',         bg,      False),
                (drv.pago_stops,     'right', '"$"#,##0.00',  bg,      False),
                (extras_t,           'center','#,##0',         bg,      False),
                (drv.pago_extra,     'right', '"$"#,##0.00',  bg,      False),
                (drv.bono,           'right', '"$"#,##0.00',  C_YELLOW if drv.bono > 0 else bg, False),
                (drv.penalizaciones, 'right', '"$"#,##0.00',  C_RED if drv.penalizaciones < 0 else bg, False),
                (drv.pago_total,     'right', '"$"#,##0.00',  C_GREEN, True),
            ]
            for ci, (val, align, fmt, rbg, bold) in enumerate(row_vals, 1):
                c = ws_p.cell(row=ri, column=ci, value=val)
                dat(c, bg=rbg, bold=bold, align=align, fmt=fmt)

        tr = len(df) + 3
        ws_p.row_dimensions[tr].height = 24
        ws_p.merge_cells(f'A{tr}:G{tr}')
        tc = ws_p[f'A{tr}']
        tc.value = "TOTALES"
        tc.font = Font(name='Arial', bold=True, color="FFFFFF", size=11)
        tc.fill = PatternFill("solid", fgColor=cc)
        tc.alignment = Alignment(horizontal='center', vertical='center')
        tc.border = bdr()
        for ci, val in [(8, df['penalizaciones'].sum()), (9, df['pago_total'].sum())]:
            c = ws_p.cell(row=tr, column=ci, value=val)
            c.font = Font(name='Arial', bold=True, color="FFFFFF", size=11)
            c.fill = PatternFill("solid", fgColor=cc)
            c.alignment = Alignment(horizontal='right', vertical='center')
            c.border = bdr()
            c.number_format = '"$"#,##0.00'

        # ── PROFIT ────────────────────────────────────────────────────────────
        ws_r = wb.create_sheet(f"📈 {code} Profit")
        ws_r.sheet_view.showGridLines = False
        cols_r = [('Driver',24),('Rutas',20),('Revenue GOFO',15),
                  ('Pago Driver',14),('Profit',13),('Margen %',11)]
        for ci, (_, w) in enumerate(cols_r, 1):
            ws_r.column_dimensions[get_column_letter(ci)].width = w

        ws_r.merge_cells(f'A1:{get_column_letter(len(cols_r))}1')
        ws_r['A1'] = f"REVENUE Y PROFIT — {code} {res['city_name']} | Revenue = gross fee driver"
        ws_r['A1'].font = Font(name='Arial', bold=True, size=12, color="FFFFFF")
        ws_r['A1'].fill = PatternFill("solid", fgColor=cc)
        ws_r['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_r.row_dimensions[1].height = 30

        ws_r.row_dimensions[2].height = 26
        for ci, (h, _) in enumerate(cols_r, 1):
            hdr(ws_r.cell(row=2, column=ci, value=h), bg=cc)

        for ri, drv in enumerate(df.itertuples(), 3):
            bg = C_STRIPE if ri % 2 == 0 else C_WHITE
            ws_r.row_dimensions[ri].height = 20
            pbg = C_GREEN if drv.profit >= 0 else C_RED
            for ci, (val, align, fmt, rbg, bold) in enumerate([
                (drv.driver,     'left',  None,          bg,  False),
                (drv.rutas,      'left',  None,          bg,  False),
                (drv.gross_fee,  'right', '"$"#,##0.00', bg,  False),
                (drv.pago_total, 'right', '"$"#,##0.00', bg,  False),
                (drv.profit,     'right', '"$"#,##0.00', pbg, True),
                (drv.margin/100, 'center','0.0%',         pbg, False),
            ], 1):
                c = ws_r.cell(row=ri, column=ci, value=val)
                dat(c, bg=rbg, bold=bold, align=align, fmt=fmt)

        tr2 = len(df) + 3
        ws_r.row_dimensions[tr2].height = 24
        tv = df['gross_fee'].sum(); tp = df['pago_total'].sum()
        tpf = res['profit']; tm2 = res['margin']
        ws_r.merge_cells(f'A{tr2}:B{tr2}')
        tc2 = ws_r[f'A{tr2}']
        tc2.value = "TOTALES"
        tc2.font = Font(name='Arial', bold=True, color="FFFFFF", size=11)
        tc2.fill = PatternFill("solid", fgColor=cc)
        tc2.alignment = Alignment(horizontal='center', vertical='center')
        tc2.border = bdr()
        for ci, val, fmt in [
            (3, res['revenue'],        '"$"#,##0.00'),
            (4, tp,                    '"$"#,##0.00'),
            (5, tpf,                   '"$"#,##0.00'),
            (6, tm2/100,               '0.0%'),
        ]:
            c = ws_r.cell(row=tr2, column=ci, value=val)
            c.font = Font(name='Arial', bold=True, color="FFFFFF", size=11)
            c.fill = PatternFill("solid", fgColor=cc)
            c.alignment = Alignment(horizontal='right' if ci<6 else 'center', vertical='center')
            c.border = bdr()
            c.number_format = fmt

    # ══ PENALIZACIONES ════════════════════════════════════════════════════════
    ws_pen = wb.create_sheet("🚨 Penalizaciones")
    ws_pen.sheet_view.showGridLines = False
    for col, w in zip('ABCDEF', [8,24,36,22,14,20]):
        ws_pen.column_dimensions[col].width = w

    ws_pen.merge_cells('A1:F1')
    ws_pen['A1'] = "PENALIZACIONES — TODAS LAS CIUDADES (descontadas del payroll del driver)"
    ws_pen['A1'].font = Font(name='Arial', bold=True, size=13, color="FFFFFF")
    ws_pen['A1'].fill = PatternFill("solid", fgColor="9C0000")
    ws_pen['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_pen.row_dimensions[1].height = 30

    ws_pen.row_dimensions[2].height = 26
    for ci, h in enumerate(['Ciudad','Driver','Waybill','Tipo','Monto','Remarks'], 1):
        hdr(ws_pen.cell(row=2, column=ci, value=h), bg="C00000")

    pen_row = 3
    for code, res in city_results.items():
        cl_df = res.get('claims_df', pd.DataFrame())
        if len(cl_df) == 0: continue
        cc = CITY_COLORS.get(code, C_ACCENT)
        for _, crow in cl_df.iterrows():
            amt = float(crow.get('amt', 0))
            bg  = C_RED if amt < 0 else C_GREEN
            ws_pen.row_dimensions[pen_row].height = 20
            for ci, (val, align, fmt) in enumerate([
                (code,                            'center', None),
                (str(crow.get('Courier', '')),    'left',   None),
                (str(crow.get('Waybill No.', '')), 'left',  None),
                (str(crow.get('Claim Type', '')), 'center', None),
                (amt,                             'right',  '"$"#,##0.00'),
                (str(crow.get('Remarks', '')),    'left',   None),
            ], 1):
                c = ws_pen.cell(row=pen_row, column=ci, value=val)
                if ci == 1:
                    c.font = Font(name='Arial', bold=True, color="FFFFFF", size=10)
                    c.fill = PatternFill("solid", fgColor=cc)
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border = bdr()
                else:
                    dat(c, bg=bg, align=align, fmt=fmt)
            pen_row += 1

    wb.save(out_path)
    return out_path

# ─── MAIN ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    files = sys.argv[1:] if len(sys.argv) > 1 else []
    if not files:
        files = sorted(glob.glob("/mnt/user-data/uploads/*-SVD-Details_of_Delivery_Fees-15_06*__1_.xlsx"))
    if not files:
        files = sorted(glob.glob("/mnt/user-data/uploads/*-SVD-Details_of_Delivery_Fees-15_06*.xlsx"))

    print(f"Procesando {len(files)} archivos...")
    city_results = {}
    for f in files:
        code = detect_city(f)
        if not code:
            print(f"  ⚠️  Saltando: {os.path.basename(f)}")
            continue
        if code in city_results:  # skip duplicates, prefer __1_ version
            continue
        print(f"  ✅ {code} — {os.path.basename(f)}")
        city_results[code] = process_file(f)

    out = f"/mnt/user-data/outputs/Reporte_GOFO_Semana25_v3_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    build_report(city_results, out)

    print(f"\n{'─'*68}")
    print(f"{'CIUDAD':<8} {'REVENUE(neto)':>14} {'CLAIMS(info)':>13} {'PAYROLL':>13} {'PROFIT':>12} {'MARGEN':>8}")
    print(f"{'─'*68}")
    totals = {'rev':0,'cl':0,'pay':0,'prof':0}
    for code, res in city_results.items():
        print(f"{code:<8} ${res['revenue']:>13,.2f} ${res['claims']:>12,.2f} ${res['total_payroll']:>12,.2f} ${res['profit']:>11,.2f} {res['margin']:>7.1f}%")
        totals['rev']+=res['revenue']; totals['cl']+=res['claims']
        totals['pay']+=res['total_payroll']; totals['prof']+=res['profit']
    print(f"{'─'*68}")
    tm=(totals['prof']/totals['rev']*100) if totals['rev']>0 else 0
    print(f"{'TOTAL':<8} ${totals['rev']:>13,.2f} ${totals['cl']:>12,.2f} ${totals['pay']:>12,.2f} ${totals['prof']:>11,.2f} {tm:>7.1f}%")
    print(f"\n✅ Reporte: {out}")

def get_zero_delivery_drivers(filepath, city_code, cfg):
    """Find drivers in Driver Summary with no deliveries this week (company loss)."""
    xl = pd.ExcelFile(filepath)
    det_raw = pd.read_excel(xl, sheet_name='Details of Delivery Fees', header=None)
    det = det_raw.iloc[1:].copy()
    det.columns = det_raw.iloc[0].values
    det = det[det['Courier'].notna()].copy()
    det['Courier'] = det['Courier'].astype(str).str.strip()
    detail_names = set(det['Courier'].unique())

    ds_raw = pd.read_excel(xl, sheet_name='Driver Summary', header=None)
    zero_drivers = []
    for _, r in ds_raw.iloc[2:].iterrows():
        name = str(r[1]).strip()
        if not name or name == 'nan': continue
        if any(x in name.lower() for x in ['payroll','gofo','profit','total']): break
        col6 = float(r[6]) if pd.notna(r[6]) else 0.0
        col3 = float(r[3]) if pd.notna(r[3]) else 0.0
        if name not in detail_names and col3 < 0:
            zero_drivers.append({
                'city': city_code, 'city_name': cfg['name'],
                'driver': name, 'rutas': '—',
                'rutas_data': [],
                'gross_fee': col6,
                'pago_stops': 0.0, 'pago_extra': 0.0, 'pago_base': 0.0,
                'bono': 0.0,
                'penalizaciones': col3,  # company absorbs this loss
                'pago_total': 0.0,       # driver gets nothing
                'company_loss': col3,    # this is YOUR loss
                'profit': col6,          # col6 is negative = pure loss
                'margin': 0.0,
                'n_penalizaciones': 1,
                'ajuste_bono': 0.0, 'ajuste_cobro': 0.0, 'ajuste_nota': 'NO TRABAJÓ — pérdida compañía',
            })
    return zero_drivers
